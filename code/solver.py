# ---------------------------------------------------------
# Tensorflow Vessel-GAN (V-GAN) Implementation
# Licensed under The MIT License [see LICENSE for details]
# Written by Cheng-Bin Jin
# ---------------------------------------------------------
import os
import time
import collections
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import tensorflow as tf
from PIL import Image

from dataset import Dataset
# noinspection PyPep8Naming
import TensorFlow_utils as tf_utils
import utils as utils
from model import CGAN
import cv2
import openpyxl

import keras.backend as K
# from sklearn.metrics import jaccard_score
# from tensorflow.keras.metrics import MeanIoU
# import tensorflow.keras.backend as K
                
class Solver(object):
    def __init__(self, flags):
        print("\n\n\n Solver Started \n\n\n")
        run_config = tf.ConfigProto()
        run_config.gpu_options.allow_growth = True
        self.sess = tf.Session(config=run_config)
 
        self.flags = flags
        self.dataset = Dataset(self.flags.dataset, self.flags)
        print('\n\n Dataset is done')
        
        self.model = CGAN(self.sess, self.flags, self.dataset.image_size,"Disc")     ## This can executed only once    ##ADDED BY ALI
        # print(self.model)
        print('\n\n CGAN is done')
        
        self.model_OC = CGAN(self.sess, self.flags, self.dataset.image_size,"Cup")     ## This can executed only once    ##ADDED BY ALI
        print('\n\n CGAN_OC is done')
        
        self.best_auc_sum = 0.
        self.best_auc_sum_OC = 0.
       
        self._make_folders()
        print('\n\n _make_folders is done \n')
        
        self.saver = tf.train.Saver()
        self.sess.run(tf.global_variables_initializer())

        tf_utils.show_all_variables()

    def _make_folders(self):
        self.model_out_dir = "./{}/train/{}/model_{}_{}/".format(self.flags.dataset, self.flags.discriminator,
                                                        self.flags.train_interval, self.flags.batch_size)
        self.model_out_dir_OC = "D:/Spider_IDE/3rd Objective/3. V_GAN  (Here)/V-GAN- (Here is my work)/codes excuted on OD + OC + classifier/{}/train/{}/model_OC_{}_{}/".format(self.flags.dataset, self.flags.discriminator,
                                                        self.flags.train_interval, self.flags.batch_size)
  
        if not os.path.isdir(self.model_out_dir):
            os.makedirs(self.model_out_dir)

        if not os.path.isdir(self.model_out_dir_OC):
                os.makedirs(self.model_out_dir_OC)    

        if self.flags.is_test:                    ## while Testing    ## Added by Ali
            self.img_out_dir = "./{}/test/{}/seg_result_{}_{}".format(self.flags.dataset,
                                                               self.flags.discriminator,
                                                               self.flags.train_interval,
                                                               self.flags.batch_size)
            
            self.img_out_dir_OC = "./{}/test/{}/seg_result_OC_{}_{}".format(self.flags.dataset,
                                                                self.flags.discriminator,
                                                                self.flags.train_interval,
                                                                self.flags.batch_size)
            
            
            self.auc_out_dir = "./{}/test/{}/auc_{}_{}".format(self.flags.dataset, self.flags.discriminator,
                                                        self.flags.train_interval, self.flags.batch_size)

            self.auc_out_dir_OC = "./{}/test/{}/auc_OC_{}_{}".format(self.flags.dataset, self.flags.discriminator,
                                                        self.flags.train_interval, self.flags.batch_size)


            if not os.path.isdir(self.img_out_dir):
                os.makedirs(self.img_out_dir)
            if not os.path.isdir(self.auc_out_dir):
                os.makedirs(self.auc_out_dir)

            if not os.path.isdir(self.img_out_dir_OC):
                os.makedirs(self.img_out_dir_OC)
            if not os.path.isdir(self.auc_out_dir_OC):
                os.makedirs(self.auc_out_dir_OC)


        elif not self.flags.is_test:             ## while Training    ## Added by Ali
            self.sample_out_dir = "D:/Spider_IDE/3rd Objective/3. V_GAN  (Here)/V-GAN- (Here is my work)/codes excuted on OD + OC + classifier/{}/train/{}/sample_{}_{}/".format(self.flags.dataset, self.flags.discriminator,
                                                              self.flags.train_interval, self.flags.batch_size)
            
            self.sample_out_dir_OC = "D:/Spider_IDE/3rd Objective/3. V_GAN  (Here)/V-GAN- (Here is my work)/codes excuted on OD + OC + classifier/{}/train/{}/sample_OC_{}_{}/".format(self.flags.dataset, self.flags.discriminator,
                                                              self.flags.train_interval, self.flags.batch_size)
         
            if not os.path.isdir(self.sample_out_dir):
                os.makedirs(self.sample_out_dir)

            if not os.path.isdir(self.sample_out_dir_OC):
                os.makedirs(self.sample_out_dir_OC)


    def train(self):
        # Create a new workbook and select the active sheet  to record d_loss and g_loss on excel
        # workbook = openpyxl.Workbook()
        # sheet = workbook.active
        # # Write the column headers
        # sheet['A1'] = 'Iteration'
        # sheet['B1'] = 'D Loss Disc'
        # sheet['C1'] = 'G Loss Disc'
        # sheet['D1'] = 'D Loss Cup'
        # sheet['E1'] = 'G Loss Cup'
        # ##################################
        
        for iter_time in range(0, self.flags.iters+1, self.flags.train_interval):
            self.sample(iter_time)  # sampling images and save them

            ''' if you need to stop on of the GANs models, you have to stop the sampling step exist in def sample() function '''

            # # train discrminator
            for iter_ in range(1, self.flags.train_interval+1):
                x_imgs, y_imgs, _ = self.dataset.train_next_batch(batch_size=self.flags.batch_size, model_type="Disc")     ## Resturns img , disc, cups      ADDED BY ALI
                d_loss = self.model.train_dis(x_imgs, y_imgs)
                self.print_info(iter_time + iter_, 'd_loss', d_loss, "Disc")


            # # train discrminator OC
            if self.flags.Use_both_GAN == True:
                for iter_ in range(1, self.flags.train_interval+1):
                    x_imgs_OC, _, y_imgs_OC = self.dataset.train_next_batch(batch_size=self.flags.batch_size, model_type="Cup")    ## Resturns img , disc, cups      ADDED BY ALI  
                    d_loss_OC = self.model_OC.train_dis_OC(x_imgs_OC, y_imgs_OC)
                    self.print_info(iter_time + iter_, 'd_loss_OC', d_loss_OC, "Cup")
  
    
            # # train generator
            for iter_ in range(1, self.flags.train_interval+1):
                x_imgs, y_imgs, _ = self.dataset.train_next_batch(batch_size=self.flags.batch_size, model_type="Disc")     ## Resturns img , disc, cups      ADDED BY ALI
                g_loss = self.model.train_gen(x_imgs, y_imgs)
                self.print_info(iter_time + iter_, 'g_loss', g_loss, "Disc")
       
                          
            # # train generator For OC
            if self.flags.Use_both_GAN == True:
                for iter_ in range(1, self.flags.train_interval+1):
                    x_imgs_OC, _, y_imgs_OC = self.dataset.train_next_batch(batch_size=self.flags.batch_size, model_type="Cup")     ## Resturns img , disc, cups      ADDED BY ALI
                    g_loss_OC = self.model_OC.train_gen_OC(x_imgs_OC, y_imgs_OC)
                    self.print_info(iter_time + iter_, 'g_loss_OC', g_loss_OC, "Cup")
    

            IOU_disc = d_loss 
            if self.flags.Use_both_GAN == True:
                IOU_cup = d_loss_OC
            
            
            auc_sum = self.eval("Disc",iter_time, IOU_disc, phase='train')
            if self.best_auc_sum < auc_sum:
                self.best_auc_sum = auc_sum
                self.save_model(iter_time,"Disc")

            if self.flags.Use_both_GAN == True:
                auc_sum_OC = self.eval("Cup", iter_time, IOU_disc, IOU_cup, phase='train')
                if self.best_auc_sum_OC < auc_sum_OC:
                    self.best_auc_sum_OC = auc_sum_OC
                    self.save_model(iter_time,"Cup")

        #     # Assuming you have variables d_loss and g_loss containing the respective values
        #     sheet.cell(row=iter_time+2, column=1).value = iter_time + 1
        #     sheet.cell(row=iter_time+2, column=2).value = d_loss     #[iter_time]
        #     sheet.cell(row=iter_time+2, column=3).value = g_loss     #[iter_time]
        #     sheet.cell(row=iter_time+2, column=4).value = d_loss_OC  #[iter_time]
        #     sheet.cell(row=iter_time+2, column=5).value = g_loss_OC  #[iter_time]

        # # Save the workbook after iteration finished
        # workbook.save('loss_values.xlsx')
        print('----------------------------------')         # To seperate between printed Losses at every iteration   ## ADDED BY ALI

            
    def test(self):
        if self.load_model():
            print(' [*] Load Success!\n')
            # Disc = self.eval("Disc",phase='test')
            # print("AUC Disc is "+str(Disc)+'\n')
            
            # print('---------------- Cup Started------------------\n ')
            # print('---------------- Cup Started------------------\n ')
            # print('---------------- Cup Started------------------\n ')            
            # print('---------------- Cup Started------------------\n ')            
            
            Cup = self.eval("Cup",phase='test')
            print("AUC Cup is "+str(Cup)+ '\n')
        else:
            print(' [!] Load Failed!\n')


        ''' Function added by Ali to draw plots, and calculate the mean/std of the PR,Recall, Acc and others'''
        ###################################################
        # mean_std_precision('Drishti-GS1/test/patch1/auc_1_1/')
        
        # mean_std_precision_OC('Drishti-GS1/test/patch1/auc_OC_1_1/')
        ###################################################
        '''#########################################'''
    

    def sample(self, iter_time):
        if np.mod(iter_time, self.flags.sample_freq) == 0:
            idx = np.random.choice(self.dataset.num_val, 2, replace=False)
            
            x_imgs, y_imgs, z_imgs = self.dataset.val_imgs[idx], self.dataset.val_Discs[idx], self.dataset.val_Cups[idx]
            
            samples = self.model.sample_imgs(x_imgs)
            if self.flags.Use_both_GAN == True:
                samples_OC = self.model_OC.sample_imgs_OC(x_imgs)     #### For OC

            # masking
            seg_samples = utils.remain_in_mask(samples, self.dataset.val_masks[idx])
            if self.flags.Use_both_GAN == True:
                seg_samples_OC = utils.remain_in_mask(samples_OC, self.dataset.val_masks[idx])    #### For OC

            # crop to original image shape
            x_imgs = utils.crop_to_original(x_imgs, self.dataset.ori_shape)
       
            seg_samples_ = utils.crop_to_original(seg_samples, self.dataset.ori_shape)
            if self.flags.Use_both_GAN == True:
                seg_samples_OC = utils.crop_to_original(seg_samples_OC, self.dataset.ori_shape)  #### For OC

            y_imgs_ = utils.crop_to_original(y_imgs, self.dataset.ori_shape)
            if self.flags.Use_both_GAN == True:
                z_imgs = utils.crop_to_original(z_imgs, self.dataset.ori_shape)  #### For OC

            # # sampling
            self.plot(x_imgs, seg_samples_, y_imgs_, iter_time, "Disc", idx=idx, save_file=self.sample_out_dir,
                      phase='train')
            
            # # sampling  OC
            if self.flags.Use_both_GAN == True:
                self.plot(x_imgs, seg_samples_OC, z_imgs, iter_time, "Cup", idx=idx, save_file=self.sample_out_dir_OC,
                          phase='train')

    def plot(self, x_imgs, samples, y_imgs, iter_time, model_type, idx=None, save_file=None, phase='train'):
        # initialize grid size
        cell_size_h, cell_size_w = self.dataset.ori_shape[0] / 100, self.dataset.ori_shape[1] / 100
            
        num_columns, margin = 3, 0.05
        width = cell_size_w * num_columns
        height = cell_size_h * x_imgs.shape[0]
        fig = plt.figure(figsize=(width, height))  # (column, row)
        gs = gridspec.GridSpec(x_imgs.shape[0], num_columns)  # (row, column)
        gs.update(wspace=margin, hspace=margin)

        # convert from normalized to original image
        x_imgs_norm = np.zeros_like(x_imgs)
        std, mean = 0., 0.
        for _ in range(x_imgs.shape[0]):
            if phase == 'train':
                std = self.dataset.val_mean_std[idx[_]]['std']
                mean = self.dataset.val_mean_std[idx[_]]['mean']
            elif phase == 'test':
                std = self.dataset.test_mean_std[idx[_]]['std']
                mean = self.dataset.test_mean_std[idx[_]]['mean']
            x_imgs_norm[_] = np.expand_dims(x_imgs[_], axis=0) * std + mean

    
        x_imgs_norm = x_imgs_norm.astype(np.uint8)


        # samples = utils.geo_shape(samples)       # Added by Ali


        # 1 channel to 3 channels
        samples_3 = np.stack((samples, samples, samples), axis=3)
        y_imgs_3 = np.stack((y_imgs, y_imgs, y_imgs), axis=3)

        imgs = [x_imgs_norm, samples_3, y_imgs_3]
        for col_index in range(len(imgs)):
            for row_index in range(x_imgs.shape[0]):
                ax = plt.subplot(gs[row_index * num_columns + col_index])
                plt.axis('off')
                plt.title(model_type)            ## Added by Ali
                ax.set_xticklabels([])
                ax.set_yticklabels([])
                ax.set_aspect('equal')
                plt.imshow(imgs[col_index][row_index].reshape(self.dataset.ori_shape[0], self.dataset.ori_shape[1], 3), cmap='Greys_r')              

        if model_type=="Disc":
            if phase == 'train':
                plt.savefig(save_file + '/{}_{}.png'.format(str(iter_time), idx[0]), bbox_inches='tight')
                plt.close(fig)
            else:
                # save compared image
                plt.savefig(os.path.join(save_file, 'compared_{}.png'.format(os.path.basename(
                    self.dataset.test_img_files[idx[0]])[:-4])), bbox_inches='tight')
                plt.close(fig)
    
                # save Disc alone, Disc should be uint8 type
                Image.fromarray(np.squeeze(samples*255).astype(np.uint8)).save(os.path.join(
                    save_file, '{}.png'.format(os.path.basename(self.dataset.test_img_files[idx[0]][:-4]))))
        else:
            if phase == 'train':
                plt.savefig(save_file + '/{}_{}_OC.png'.format(str(iter_time), idx[0]), bbox_inches='tight')
                plt.close(fig)
            else:
                # save compared image
                plt.savefig(os.path.join(save_file, 'compared_OC_{}.png'.format(os.path.basename(
                    self.dataset.test_img_files[idx[0]])[:-4])), bbox_inches='tight')
                plt.close(fig)
    
                # save Cup alone, Cup should be uint8 type
                Image.fromarray(np.squeeze(samples*255).astype(np.uint8)).save(os.path.join(
                    save_file, '{}_OC.png'.format(os.path.basename(self.dataset.test_img_files[idx[0]][:-4]))))                

            

    def print_info(self, iter_time, name, loss,model_type):
        if np.mod(iter_time, self.flags.print_freq) == 0:
            if model_type=='Disc':
                ord_output = collections.OrderedDict([(name, loss), ('dataset', self.flags.dataset),
                                                      ('discriminator', self.flags.discriminator),
                                                      ('train_interval', np.float32(self.flags.train_interval)),
                                                      ('gpu_index', self.flags.gpu_index)])
                utils.print_metrics(iter_time, ord_output)
            else:
                ord_output = collections.OrderedDict([(name, loss), ('dataset', self.flags.dataset),
                                                      ('discriminator_OC', self.flags.discriminator_OC),
                                                      ('train_interval_OC', np.float32(self.flags.train_interval_OC)),
                                                      ('gpu_index', self.flags.gpu_index)])
                utils.print_metrics_OC(iter_time, ord_output)                

    def eval(self, model_type, iter_time=0, IOU_disc=0, IOU_cup=0, phase='train'):
        total_time, auc_sum = 0., 0.
        if np.mod(iter_time, self.flags.eval_freq) == 0:
            num_data, imgs, Discs, masks = None, None, None, None
            if model_type=='Disc':
                if phase == 'train':
                    num_data = self.dataset.num_val
                    imgs = self.dataset.val_imgs
                    Discs = self.dataset.val_Discs
                    masks = self.dataset.val_masks
                elif phase == 'test':
                    num_data = self.dataset.num_test
                    imgs = self.dataset.test_imgs
                    Discs = self.dataset.test_Discs
                    masks = self.dataset.test_masks
            else:
                if phase == 'train':
                    num_data = self.dataset.num_val
                    imgs = self.dataset.val_imgs    ## Original  # imgs = self.dataset.val_imgs
                    Discs = self.dataset.val_Cups    ## Original  # Discs = self.dataset.val_Discs
                    masks = self.dataset.val_masks
                elif phase == 'test':
                    num_data = self.dataset.num_test
                    imgs = self.dataset.test_imgs
                    Discs = self.dataset.test_Cups
                    masks = self.dataset.test_masks


            Discs2 = Discs    ## Added by ALI for testing single image a time
            masks2 = masks    ## Added by ALI for testing single image a time
            
            test_one_by_one = False    ## Added by Ali
            PGS_avtivated = False
            
            generated = []
            for iter_ in range(num_data):
                x_img = imgs[iter_]
                x_img2 =x_img  ## Added by Ali
                x_img = np.expand_dims(x_img, axis=0)  # (H, W, C) to (1, H, W, C)

                
                # measure inference time
                start_time = time.time()
                
                if phase == 'train':      ## we seperate these two line of codes into test and train to apply our function on the testing only not on the training
                    if model_type=='Disc':
                        generated_Disc = self.model.sample_imgs(x_img)
                        ## if self.best_auc_sum >=1.90:
                        ##     generated_Disc = utils.geo_shape(generated_Disc)    ## Added by Ali
                    else:
                        generated_Disc = self.model_OC.sample_imgs_OC(x_img)
                        ## if self.best_auc_sum_OC >=1.90:
                        ##     generated_Disc = utils.geo_shape(generated_Disc)    ## Added by Ali
                
                else: #if phase == 'test':
                    if model_type=='Disc':
                        generated_Disc = self.model.sample_imgs(x_img)   ## Shape is (1, 720, 720, 1)
                        if PGS_avtivated == True:
                            generated_Disc,Ver_diameter = utils.geo_shape(generated_Disc)    ## Added by Ali
                            ## print("Ver_diameter_disc : " + str(Ver_diameter) +'\n')
                    else:
                        generated_Disc = self.model_OC.sample_imgs_OC(x_img)
                        if PGS_avtivated == True:
                             generated_Disc,Ver_diameter = utils.geo_shape(generated_Disc)    ## Added by Ali
                             ## print("Ver_diameter_cup : " + str(Ver_diameter) +'\n')
                
                
                #################################################
                #### ADDED BY ALI to measure the accuracy of segmentating each image
                #### calculate measurements    
                if test_one_by_one== True and self.flags.is_test ==True:
                    Discs = Discs2[iter_]  ## Added by ALI for testing single image a time
                    masks = masks2[iter_]  ## Added by ALI for testing single image a time
    
                    total_time += (time.time() - start_time)
    
                    generated = []
                    print ('--------- image' + str(iter_) + '------------------ \n' )
                    # print("Ver_diameter : " + str(Ver_diameter) +'\n')
                    
                    x= np.squeeze(generated_Disc, axis=(0, 3))
                    generated = np.asarray(x)        ## shape is (720, 720)
                    
                    fig, ax = plt.subplots(1, 2, figsize=(10, 5))
                    ax[0].imshow(generated, cmap='gray')
                    ax[0].set_title('generated '+str(iter_))
                    ax[1].imshow(Discs, cmap='gray')
                    ax[1].set_title('Discs ' + str(iter_))
                    plt.show()
                                                              
                    #### calculate measurements
                    if model_type=='Disc':
                        auc_sum = self.measure(generated, Discs, masks, num_data, iter_time, phase, IOU_disc, total_time,"Disc")
                    else:
                        auc_sum = self.measure(generated, Discs, masks, num_data, iter_time, phase, IOU_cup, total_time,"Cup")
                    #############   End of testing single image one by one         ADDED BY ALI
                    ##############################
                
                if test_one_by_one== False or self.flags.is_test == False:
        
            ####### ORIGINAL
            ######## Just reactivate these lines for full testing of the whole test set                 
                    total_time += (time.time() - start_time)
    
                    generated.append(np.squeeze(generated_Disc, axis=(0, 3)))  # (1, H, W, 1) to (H, W)
                    
            if test_one_by_one== False or self.flags.is_test == False:
                generated = np.asarray(generated)
              
                # calculate measurements
                if model_type=='Disc':
                    auc_sum = self.measure(generated, Discs, masks, num_data, iter_time, phase, IOU_disc, total_time,"Disc")
                else:
                    auc_sum = self.measure(generated, Discs, masks, num_data, iter_time, phase, IOU_cup, total_time,"Cup")
    
                if phase == 'test':
                    # save test images
                    segmented_Disc = utils.remain_in_mask(generated, masks)
    
                    # crop to original image shape
                    if model_type=='Disc':
                        imgs_ = utils.crop_to_original(imgs, self.dataset.ori_shape)
                        cropped_Disc = utils.crop_to_original(segmented_Disc, self.dataset.ori_shape)
                        Discs_ = utils.crop_to_original(Discs, self.dataset.ori_shape)
                        
                        for idx in range(num_data):
                            self.plot(np.expand_dims(imgs_[idx], axis=0),
                                      np.expand_dims(cropped_Disc[idx], axis=0),
                                      np.expand_dims(Discs_[idx], axis=0),
                                      'test', "Disc", idx=[idx], save_file=self.img_out_dir, phase='test')
                    
                    else:
                        imgs_ = utils.crop_to_original(imgs, self.dataset.ori_shape)
                        cropped_Disc = utils.crop_to_original(segmented_Disc, self.dataset.ori_shape)
                        Discs_ = utils.crop_to_original(Discs, self.dataset.ori_shape)
    
                        for idx in range(num_data):
                            self.plot(np.expand_dims(imgs_[idx], axis=0),
                                      np.expand_dims(cropped_Disc[idx], axis=0),
                                      np.expand_dims(Discs_[idx], axis=0),
                                      'test', "Cup", idx=[idx], save_file=self.img_out_dir_OC, phase='test')

              
        return auc_sum
  
    def measure(self, generated, Discs, masks, num_data, iter_time, phase, IOU, total_time, model_type):
        # masking
        Discs_in_mask, generated_in_mask = utils.pixel_values_in_mask(Discs, generated, masks)

        ############    ADDED BY ALI        
        if phase == 'test':     # only in testing coz activating this measurment during training raises an error
            '''Clacluating IOU'''     
            ##### Cast y_true and y_pred to float32
            y_true = K.cast(Discs_in_mask, dtype='float32')
            y_pred = K.cast(generated_in_mask, dtype='float32')
            ##### Flatten the predictions and true labels
            y_true_flat = K.flatten(y_true)
            y_pred_flat = K.flatten(y_pred)
            ##### Calculate the intersection between the two label sets
            intersection = K.sum(y_true_flat * y_pred_flat)
            ##### Calculate the union between the two label sets
            union = K.sum(y_true_flat) + K.sum(y_pred_flat) - intersection
            ##### Avoid division by zero
            iou_score = K.switch(K.equal(union, 0), K.zeros_like(union), intersection / K.cast(union, dtype='float32'))
            #### Evaluate the IOU score using a Keras session   
            with K.get_session().as_default():
                iou_score_value = iou_score.eval()
            print("IOU : " + str(iou_score_value) +'\n')
            '''# #####################################'''
        ############    ADDED BY ALI        
        
        # averaging processing time
        avg_pt = (total_time / num_data) * 1000  # average processing tiem

        # evaluate Area Under the Curve of ROC and Precision-Recall
        auc_roc = utils.AUC_ROC(Discs_in_mask, generated_in_mask)
        auc_pr = utils.AUC_PR(Discs_in_mask, generated_in_mask)

        # binarize to calculate Dice Coeffient
        binarys_in_mask = utils.threshold_by_otsu(generated, masks)
        dice_coeff = utils.dice_coefficient_in_train(Discs_in_mask, binarys_in_mask)
        acc, sensitivity, specificity = utils.misc_measures(Discs_in_mask, binarys_in_mask)
        score = auc_pr + auc_roc + dice_coeff + acc + sensitivity + specificity

        # auc_sum for saving best model in training
        auc_sum = auc_roc + auc_pr

        # print information
        if model_type == "Disc":
            ord_output = collections.OrderedDict([('auc_pr', auc_pr), ('auc_roc', auc_roc),
                                                  ('dice_coeff', dice_coeff), ('acc', acc),
                                                  ('sensitivity', sensitivity), ('specificity', specificity),
                                                  ('IOU_disc', IOU),
                                                  ('score', score), ('auc_sum', auc_sum),
                                                  ('best_auc_sum', self.best_auc_sum), ('avg_pt', avg_pt)])
            utils.print_metrics(iter_time, ord_output)

        else:
            ord_output = collections.OrderedDict([('auc_pr_OC', auc_pr), ('auc_roc_OC', auc_roc),
                                                  ('dice_coeff_OC', dice_coeff), ('acc_OC', acc),
                                                  ('sensitivity_OC', sensitivity), ('specificity_OC', specificity),
                                                  ('IOU_cup', IOU),
                                                  ('score_OC', score), ('auc_sum_OC', auc_sum),
                                                  ('best_auc_sum_OC', self.best_auc_sum_OC), ('avg_pt_OC', avg_pt)])
            utils.print_metrics_OC(iter_time, ord_output)

        # write in tensorboard when in train mode only
        if phase == 'train':
            if model_type == "Disc":
                self.model.measure_assign(
                    auc_pr, auc_roc, dice_coeff, acc, sensitivity, specificity, IOU, score, iter_time,"Disc")
            else:
                self.model_OC.measure_assign(
                    auc_pr, auc_roc, dice_coeff, acc, sensitivity, specificity, IOU, score, iter_time,"Cup")
                
        elif phase == 'test':    
            # write in npy format for evaluation
            if model_type == "Disc":
                utils.save_obj(Discs_in_mask, generated_in_mask,
                               os.path.join(self.auc_out_dir, "auc_roc.npy"),
                               os.path.join(self.auc_out_dir, "auc_pr.npy"),"Disc")
            else:
                utils.save_obj(Discs_in_mask, generated_in_mask,
                               os.path.join(self.auc_out_dir_OC, "auc_roc_OC.npy"),
                               os.path.join(self.auc_out_dir_OC, "auc_pr_OC.npy"),"Cup")

        return auc_sum

    def save_model(self, iter_time, model_type):
        if model_type == "Disc":
            self.model.best_auc_sum_assign(self.best_auc_sum)
    
            model_name = "iter_{}_{}_auc_sum_{:.3}".format(iter_time,self.flags.discriminator, self.best_auc_sum)
            
            print('\n\n self.model_out_dir is : '+ str(self.model_out_dir))
            
            self.saver.save(self.sess, os.path.join(self.model_out_dir, model_name))
    
            print('===================================================')
            print('                     Model saved for Disc!                  ')
            print(' Best auc_sum: {:.3}'.format(self.best_auc_sum))
            print('===================================================\n')
        
        else:
            self.model_OC.best_auc_sum_assign_OC(self.best_auc_sum_OC)
    
            model_name = "iter_{}_{}_auc_sum_OC_{:.3}".format(iter_time,self.flags.discriminator_OC, self.best_auc_sum_OC)
            
            print('\n\n self.model_out_dir_OC is : '+ str(self.model_out_dir_OC))
            
            self.saver.save(self.sess, os.path.join(self.model_out_dir_OC, model_name))
    
            print('===================================================')
            print('                     Model saved For Cup!                  ')
            print(' Best auc_sum_OC: {:.3}'.format(self.best_auc_sum_OC))
            print('===================================================\n')
                    

    def load_model(self):
        flag= 0
        flag2=0
        print(' [*] Reading checkpoint...')

        ##### for Disc saved model
        ckpt = tf.train.get_checkpoint_state(self.model_out_dir)
        if ckpt and ckpt.model_checkpoint_path:
            ckpt_name = os.path.basename(ckpt.model_checkpoint_path)
            self.saver.restore(self.sess, os.path.join(self.model_out_dir, ckpt_name))

            self.best_auc_sum = self.sess.run(self.model.best_auc_sum)
            print('====================================================')
            print('                     Model saved!                   ')
            print(' Best auc_sum: {:.3}'.format(self.best_auc_sum))
            print('====================================================')
            flag = 1
        
        ##### for CUP saved model
        ckpt_OC = tf.train.get_checkpoint_state(self.model_out_dir_OC)
        if ckpt_OC and ckpt_OC.model_checkpoint_path:
            ckpt_name_OC = os.path.basename(ckpt_OC.model_checkpoint_path)
            self.saver.restore(self.sess, os.path.join(self.model_out_dir_OC, ckpt_name_OC))

            self.best_auc_sum_OC = self.sess.run(self.model_OC.best_auc_sum_OC)
            print('====================================================')
            print('                     Model saved!                   ')
            print(' Best auc_sum: {:.3}'.format(self.best_auc_sum_OC))
            print('====================================================')
            flag2 = 1
        
        # print('flag : ' + str(flag))
        # print('\n\n flag2 : ' + str(flag2))
        
        if flag == 1 or flag2 == 1:
            # print("\n\n Trueeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee \n ")
            return True
        else:
            # print("\n\n Fasdasdasdasdasas \n ")
            return False

















